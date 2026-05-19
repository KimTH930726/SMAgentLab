"""파일 포맷별 파싱 어댑터 — 다양한 입력을 통일 포맷으로 변환."""
import csv
import io
import logging
import re
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)


@dataclass
class ParsedDocument:
    """파싱된 문서 통일 포맷."""
    source_type: str  # txt, md, pdf, xlsx, csv, confluence, web
    source_name: str
    raw_text: str
    sections: list[dict] = field(default_factory=list)  # [{title, content, level}]
    tables: list[dict] = field(default_factory=list)     # [{headers, rows}]
    metadata: dict = field(default_factory=dict)


def parse_text(content: str, filename: str) -> ParsedDocument:
    """일반 텍스트 파싱."""
    return ParsedDocument(
        source_type="txt",
        source_name=filename,
        raw_text=content,
    )


def parse_markdown(content: str, filename: str) -> ParsedDocument:
    """마크다운 파싱 — 헤더 기반 섹션 추출."""
    sections: list[dict] = []
    current_title = ""
    current_level = 0
    current_lines: list[str] = []

    for line in content.split("\n"):
        m = re.match(r'^(#{1,4})\s+(.+)', line)
        if m:
            # 이전 섹션 저장
            if current_lines or current_title:
                sections.append({
                    "title": current_title,
                    "content": "\n".join(current_lines).strip(),
                    "level": current_level,
                })
            current_level = len(m.group(1))
            current_title = m.group(2).strip()
            current_lines = []
        else:
            current_lines.append(line)

    # 마지막 섹션
    if current_lines or current_title:
        sections.append({
            "title": current_title,
            "content": "\n".join(current_lines).strip(),
            "level": current_level,
        })

    # 테이블 추출 (마크다운 테이블)
    tables = _extract_md_tables(content)

    return ParsedDocument(
        source_type="md",
        source_name=filename,
        raw_text=content,
        sections=sections,
        tables=tables,
    )


def parse_pdf(content_bytes: bytes, filename: str) -> ParsedDocument:
    """PDF 파싱 — pymupdf 사용."""
    try:
        import pymupdf  # PyMuPDF
    except ImportError:
        try:
            import fitz as pymupdf  # 구 버전 호환
        except ImportError:
            logger.warning("pymupdf 미설치 — PDF를 텍스트로만 추출합니다.")
            # fallback: 바이너리를 디코딩 시도 (당연히 실패하지만 에러 메시지용)
            raise ImportError("PDF 파싱을 위해 pymupdf를 설치하세요: pip install pymupdf")

    doc = pymupdf.open(stream=content_bytes, filetype="pdf")

    # 암호화/비밀번호 보호 PDF 감지 — needs_pass 또는 is_encrypted 속성 확인
    if getattr(doc, "needs_pass", False) or getattr(doc, "is_encrypted", False):
        raise ValueError(
            "암호화(비밀번호 보호)된 PDF는 등록할 수 없습니다. "
            "원본 PDF의 비밀번호를 해제한 후 다시 업로드해주세요. "
            "(Acrobat → 도구 → 보호 → 암호화 → 보안 제거)"
        )

    pages: list[str] = []
    all_text_lines: list[str] = []

    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text()
        pages.append(text)
        all_text_lines.append(f"--- Page {page_num + 1} ---")
        all_text_lines.append(text)

    raw_text = "\n".join(all_text_lines)

    # 헤더 패턴으로 섹션 추출 시도 (PDF에서 추출된 텍스트 기반)
    sections = _extract_sections_from_text(raw_text)

    return ParsedDocument(
        source_type="pdf",
        source_name=filename,
        raw_text=raw_text,
        sections=sections,
        metadata={"page_count": len(doc)},
    )


def parse_xlsx(content_bytes: bytes, filename: str) -> ParsedDocument:
    """Excel(.xlsx) 파싱 — openpyxl 사용. 시트별로 텍스트 변환."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("XLSX 파싱을 위해 openpyxl을 설치하세요: pip install openpyxl")

    wb = load_workbook(io.BytesIO(content_bytes), data_only=True, read_only=True)

    sections: list[dict] = []
    tables: list[dict] = []
    all_text_lines: list[str] = []
    total_rows = 0

    for sheet in wb.worksheets:
        sheet_name = sheet.title
        rows_data: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            # 빈 행은 스킵
            if not any(c is not None and str(c).strip() for c in row):
                continue
            cells = [("" if c is None else str(c).strip()) for c in row]
            rows_data.append(cells)

        if not rows_data:
            continue

        # 첫 행을 헤더로 가정
        headers = rows_data[0]
        body_rows = rows_data[1:] if len(rows_data) > 1 else []
        total_rows += len(body_rows)

        # 테이블 메타 보관
        tables.append({"sheet": sheet_name, "headers": headers, "rows": body_rows})

        # 텍스트 직렬화: "헤더: 값" per row (RAG 검색에 유리)
        sheet_text_lines = [f"## 시트: {sheet_name}"]
        for row in body_rows:
            row_parts = []
            for h, v in zip(headers, row):
                if v:
                    row_parts.append(f"{h}: {v}" if h else v)
            if row_parts:
                sheet_text_lines.append(" | ".join(row_parts))

        section_text = "\n".join(sheet_text_lines[1:])
        all_text_lines.extend(sheet_text_lines)
        all_text_lines.append("")

        sections.append({
            "title": f"시트: {sheet_name}",
            "content": section_text,
            "level": 1,
        })

    raw_text = "\n".join(all_text_lines).strip()

    return ParsedDocument(
        source_type="xlsx",
        source_name=filename,
        raw_text=raw_text,
        sections=sections,
        tables=tables,
        metadata={"sheet_count": len(wb.worksheets), "total_rows": total_rows},
    )


def parse_csv(content_bytes: bytes, filename: str) -> ParsedDocument:
    """CSV 파싱 — 첫 행 헤더로 가정, 행별 직렬화."""
    text = content_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any(c.strip() for c in r)]
    if not rows:
        return ParsedDocument(source_type="csv", source_name=filename, raw_text="")

    headers = [h.strip() for h in rows[0]]
    body_rows = rows[1:]

    lines = [f"## {filename}"]
    for row in body_rows:
        row_parts = []
        for h, v in zip(headers, row):
            v = v.strip()
            if v:
                row_parts.append(f"{h}: {v}" if h else v)
        if row_parts:
            lines.append(" | ".join(row_parts))

    return ParsedDocument(
        source_type="csv",
        source_name=filename,
        raw_text="\n".join(lines),
        tables=[{"headers": headers, "rows": body_rows}],
        metadata={"row_count": len(body_rows)},
    )


def parse_file(content_bytes: bytes, filename: str) -> ParsedDocument:
    """파일 확장자로 적절한 어댑터 선택."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "pdf":
        return parse_pdf(content_bytes, filename)
    elif ext in ("xlsx", "xlsm"):
        return parse_xlsx(content_bytes, filename)
    elif ext == "csv":
        return parse_csv(content_bytes, filename)
    elif ext in ("md", "markdown"):
        return parse_markdown(content_bytes.decode("utf-8-sig"), filename)
    elif ext in ("txt", "log", "text"):
        return parse_text(content_bytes.decode("utf-8-sig"), filename)
    else:
        # fallback: 텍스트로 시도
        try:
            text = content_bytes.decode("utf-8-sig")
            return parse_text(text, filename)
        except UnicodeDecodeError:
            raise ValueError(f"지원하지 않는 파일 형식: {ext}")


# ── 내부 헬퍼 ────────────────────────────────────────────────────────────────

def _extract_sections_from_text(text: str) -> list[dict]:
    """텍스트에서 번호 매기기 패턴 (1. 2. 또는 # 헤더) 기반 섹션 추출."""
    sections: list[dict] = []
    # 패턴: "1. ", "1-1. ", "## " 등
    pattern = re.compile(r'^(?:#{1,4}\s+|(?:\d+[\.\-])+\s*)', re.MULTILINE)

    parts = pattern.split(text)
    titles = pattern.findall(text)

    # 첫 부분 (헤더 없는 서두)
    if parts and parts[0].strip():
        sections.append({"title": "", "content": parts[0].strip(), "level": 0})

    for i, title in enumerate(titles):
        content = parts[i + 1].strip() if i + 1 < len(parts) else ""
        level = title.count("#") if "#" in title else 1
        sections.append({
            "title": title.strip().rstrip("."),
            "content": content,
            "level": level,
        })

    return sections


def _extract_md_tables(text: str) -> list[dict]:
    """마크다운 테이블 추출."""
    tables: list[dict] = []
    lines = text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        # 테이블 헤더 감지: | col1 | col2 |
        if line.startswith("|") and line.endswith("|") and line.count("|") >= 3:
            headers = [h.strip() for h in line.split("|")[1:-1]]
            # 다음 줄이 구분선인지 확인: |---|---|
            if i + 1 < len(lines) and re.match(r'^\|[\s\-:]+\|', lines[i + 1].strip()):
                rows: list[list[str]] = []
                j = i + 2
                while j < len(lines):
                    row_line = lines[j].strip()
                    if not row_line.startswith("|"):
                        break
                    cells = [c.strip() for c in row_line.split("|")[1:-1]]
                    rows.append(cells)
                    j += 1
                if rows:
                    tables.append({"headers": headers, "rows": rows})
                i = j
                continue
        i += 1
    return tables
