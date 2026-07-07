"""Text2SQL 스키마 엑셀 임포터 — 헤더 자동 매핑 + 유효성 검사."""
from __future__ import annotations

import io
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

# ── 헤더 후보 매핑 테이블 ──────────────────────────────────────────────────────

_HEADER_CANDIDATES: dict[str, list[str]] = {
    "table_name": [
        "table_name", "tablename", "table", "테이블명", "테이블", "테이블이름",
        "table name", "tbl", "tbl_name",
    ],
    "column_name": [
        "column_name", "columnname", "column", "col", "col_name",
        "컬럼명", "컬럼", "컬럼이름", "열이름", "열명", "속성명", "attribute",
        "column name",
    ],
    "data_type": [
        "data_type", "datatype", "type", "dtype", "col_type",
        "데이터타입", "타입", "자료형", "데이터형", "유형",
        "data type",
    ],
    "is_pk": [
        "is_pk", "pk", "primary_key", "primarykey", "primary key",
        "기본키", "pk여부", "기본 키", "주키",
    ],
    "fk_reference": [
        "fk_reference", "fk", "foreign_key", "foreignkey", "foreign key",
        "외래키", "fk참조", "참조", "reference", "ref",
    ],
    "description": [
        "description", "desc", "comment", "remarks", "remark", "note",
        "설명", "비고", "주석", "코멘트", "설명문",
    ],
}

REQUIRED_FIELDS = {"table_name", "column_name", "data_type"}
OPTIONAL_FIELDS = {"is_pk", "fk_reference", "description"}
ALL_FIELDS = REQUIRED_FIELDS | OPTIONAL_FIELDS


def _normalize(s: str) -> str:
    return s.strip().lower().replace(" ", "_").replace("-", "_")


def map_headers(raw_headers: list[str]) -> dict[str, Optional[str]]:
    """엑셀 헤더 → 필드명 매핑.

    Returns:
        {field: original_header or None}  — None이면 매핑 실패
    """
    normalized = {_normalize(h): h for h in raw_headers if h}
    result: dict[str, Optional[str]] = {f: None for f in ALL_FIELDS}

    for field, candidates in _HEADER_CANDIDATES.items():
        for cand in candidates:
            key = _normalize(cand)
            if key in normalized:
                result[field] = normalized[key]
                break

    return result


def _parse_bool(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().upper()
    return s in ("Y", "YES", "TRUE", "1", "PK", "O", "V", "✓", "●")


def parse_excel(file_bytes: bytes) -> dict:
    """엑셀 바이트 → 파싱 결과.

    Returns:
        {
            "header_mapping": {field: original_col | None},
            "missing_required": [field, ...],
            "rows": [{"table_name":..., "column_name":..., ...}, ...],
            "warnings": ["중복 (table.col) 2건", ...],
            "error": str | None,
        }
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return {"error": f"엑셀 파일을 읽을 수 없습니다: {e}", "rows": [], "header_mapping": {}, "missing_required": [], "warnings": []}

    if not raw_rows:
        return {"error": "파일이 비어 있습니다.", "rows": [], "header_mapping": {}, "missing_required": [], "warnings": []}

    # 헤더 행 찾기 (첫 번째 비어있지 않은 행)
    header_row_idx = 0
    for i, row in enumerate(raw_rows):
        if any(cell is not None for cell in row):
            header_row_idx = i
            break

    raw_headers = [str(c).strip() if c is not None else "" for c in raw_rows[header_row_idx]]
    mapping = map_headers(raw_headers)

    missing_required = [f for f in REQUIRED_FIELDS if mapping[f] is None]
    if missing_required:
        discovered = [h for h in raw_headers if h]
        return {
            "error": (
                f"필수 컬럼을 찾을 수 없습니다.\n"
                f"필요: {', '.join(missing_required)}\n"
                f"발견된 헤더: {discovered}"
            ),
            "rows": [],
            "header_mapping": mapping,
            "missing_required": missing_required,
            "warnings": [],
        }

    # 헤더 → 인덱스 맵
    col_idx: dict[str, int] = {}
    for field, orig_header in mapping.items():
        if orig_header and orig_header in raw_headers:
            col_idx[field] = raw_headers.index(orig_header)

    # 데이터 파싱
    rows = []
    seen: set[tuple[str, str]] = set()
    warnings: list[str] = []

    for raw in raw_rows[header_row_idx + 1:]:
        def get(field: str):
            idx = col_idx.get(field)
            if idx is None or idx >= len(raw):
                return None
            v = raw[idx]
            return str(v).strip() if v is not None else None

        table = get("table_name")
        column = get("column_name")
        dtype = get("data_type")

        # 완전히 빈 행 스킵
        if not table and not column:
            continue

        if not table:
            warnings.append(f"table_name 누락 행 스킵: column='{column}'")
            continue
        if not column:
            warnings.append(f"column_name 누락 행 스킵: table='{table}'")
            continue

        dtype = dtype or "UNKNOWN"

        is_pk_raw = get("is_pk")
        is_pk = _parse_bool(is_pk_raw)

        fk_reference = get("fk_reference") or None
        description = get("description") or ""

        key = (table.upper(), column.upper())
        if key in seen:
            warnings.append(f"중복 항목 스킵: {table}.{column}")
            continue
        seen.add(key)

        rows.append({
            "table_name": table,
            "column_name": column,
            "data_type": dtype,
            "is_pk": is_pk,
            "fk_reference": fk_reference,
            "description": description,
        })

    if not rows:
        return {
            "error": "파싱된 데이터 행이 없습니다. 헤더 다음 행부터 데이터가 있어야 합니다.",
            "rows": [],
            "header_mapping": mapping,
            "missing_required": [],
            "warnings": warnings,
        }

    return {
        "error": None,
        "rows": rows,
        "header_mapping": mapping,
        "missing_required": [],
        "warnings": warnings,
    }


_sample_workbook_cache: Optional[bytes] = None


def build_sample_workbook() -> bytes:
    """샘플 템플릿 엑셀(.xlsx) 바이트 생성 — 권장 헤더 + 예시 데이터 4행.

    헤더는 _HEADER_CANDIDATES(파서가 인식하는 필드 목록)에서 그대로 파생시켜
    파서와 템플릿이 어긋나지 않게 한다. 출력은 항상 동일하므로 최초 호출 시 캐싱한다.
    """
    global _sample_workbook_cache
    if _sample_workbook_cache is not None:
        return _sample_workbook_cache

    headers = list(_HEADER_CANDIDATES.keys())
    sample_rows = [
        ["USERS", "ID", "BIGINT", "Y", "", "사용자 고유 식별자"],
        ["USERS", "EMAIL", "VARCHAR(255)", "N", "", "사용자 이메일"],
        ["ORDERS", "ID", "BIGINT", "Y", "", "주문 고유 식별자"],
        ["ORDERS", "USER_ID", "BIGINT", "N", "USERS.ID", "주문한 사용자 ID"],
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "schema"
    ws.append(headers)
    for row in sample_rows:
        ws.append(row)

    for col_idx, header in enumerate(headers, start=1):
        width = max(len(header) + 2, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    _sample_workbook_cache = buf.getvalue()
    return _sample_workbook_cache


def rows_to_tables(rows: list[dict]) -> list[dict]:
    """파싱된 행 목록 → get_tables() 규격으로 변환.

    Returns:
        [{"table_name": str, "columns": [{"name", "type", "is_pk", "fk_reference"}, ...]}, ...]
    """
    tables: dict[str, list[dict]] = {}
    for row in rows:
        tname = row["table_name"]
        if tname not in tables:
            tables[tname] = []
        tables[tname].append({
            "name": row["column_name"],
            "type": row["data_type"],
            "is_pk": row["is_pk"],
            "fk_reference": row.get("fk_reference"),
            "description": row.get("description", ""),
        })

    return [{"table_name": t, "columns": cols} for t, cols in tables.items()]
